import ezdxf
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string



def read_lwpolylines_from_dxf(file_path):
    """
    Reads all LWPOLYLINE entities from a DXF file provided as file_path.

    :param file_path: Path to the DXF file
    :return: A list of dictionaries containing information about each LWPOLYLINE
    """
    try:
        # Load the DXF document
        doc = ezdxf.readfile(file_path)
        # Access the modelspace where entities are typically stored
        msp = doc.modelspace()
        
        # Initialize a list to hold information about all LWPOLYLINEs
        lwpolyline_data = []
        
        # Iterate over all LWPOLYLINE entities in the modelspace
        for lwpolyline in msp.query("LWPOLYLINE"):
            # Extract vertex data as tuples of (x, y, start_width, end_width, bulge)
            points = [(float(x), float(y)) for x, y, _, _, _ in lwpolyline]
            
            # Store the extracted information
            data = {
                "points": points,  # List of vertex tuples
                "is_closed": lwpolyline.is_closed,  # Whether the polyline is closed
                "layer": lwpolyline.dxf.layer,  # Layer of the polyline
            }
            lwpolyline_data.append(data)
        
        return lwpolyline_data
    except IOError:
        print("Could not read the DXF file. Please check the file path.")
        return []
    except ezdxf.DXFStructureError:
        print("Invalid DXF file structure.")
        return []


def validate_rectangle(polyline):
    """
    Validates if a polyline forms a rectangle by checking dot products of consecutive vectors.
    
    :param polyline: Dictionary containing points of the polyline.
    :return: True if it is a rectangle, False otherwise.
    """
    points = polyline["points"]  # List of (x, y) tuples

    # Ensure the polyline has exactly 4 vertices and is closed
    if len(points) != 4 or not polyline["is_closed"]:
        return False

    # Calculate vectors for consecutive points (ensure loop closure with points[0])
    vectors = [(points[i][0] - points[i - 1][0], points[i][1] - points[i - 1][1]) for i in range(4)]

    # Check dot products of consecutive vectors
    for i in range(4):
        v1 = vectors[i]
        v2 = vectors[(i + 1) % 4]  # Wrap around to check last with first
        dot_product = v1[0] * v2[0] + v1[1] * v2[1]
        
        # If the dot product is not zero, it's not a rectangle
        if dot_product != 0:
            return False

    return True



def check_rectangle_properties(polyline):
    """
    Checks if a polyline rectangle is horizontal or vertical, finds its center coordinate,
    and calculates its average height and width.
    
    :param polyline: Dictionary containing points of the polyline.
    :return: A tuple (orientation, center, average_height, average_width) if valid, otherwise None.
             orientation: 'Horizontal' or 'Vertical'
             center: (x_center, y_center)
             average_height: The height of the rectangle
             average_width: The width of the rectangle
    """
    points = polyline["points"]  # List of (x, y) tuples

    # Ensure the polyline has exactly 4 vertices and is closed
    if len(points) != 4 or not polyline["is_closed"]:
        return None  # Not a valid rectangle
    
    # Find the min and max x and y coordinates
    min_x = min(point[0] for point in points)
    max_x = max(point[0] for point in points)
    min_y = min(point[1] for point in points)
    max_y = max(point[1] for point in points)
    
    # Calculate the differences
    x_diff = max_x - min_x
    y_diff = max_y - min_y

    # Calculate the center coordinate
    x_center = (min_x + max_x) / 2
    y_center = (min_y + max_y) / 2
    center = (x_center, y_center)

    # Calculate height and width
    average_height = y_diff
    average_width = x_diff

    # Determine orientation
    if y_diff > x_diff:
        orientation = "Vertical"
        cellcenter = (x_center, (max_y - ((max_y - y_center) / 2)))
    else:
        orientation = "Horizontal"
        cellcenter = ((max_x - ((max_x - x_center) / 2)), y_center)

    return orientation, center, average_height, average_width, cellcenter


# Function to move all polylines to the origin
def move_polylines_to_origin(lwpolylines):
    """
    Moves all polylines to the origin by shifting all points to account for the minimum x and y.
    """
    global_min_x = float('inf')
    global_min_y = float('inf')

    for polyline in lwpolylines:
        for point in polyline["points"]:
            global_min_x = min(global_min_x, point[0])
            global_min_y = min(global_min_y, point[1])

    for polyline in lwpolylines:
        updated_points = []
        for point in polyline["points"]:
            updated_x = point[0] - global_min_x
            updated_y = point[1] - global_min_y
            updated_points.append((updated_x, updated_y))
        polyline["points"] = updated_points

    return lwpolylines

# Function to mirror points across the x-axis
def mirror_points_across_x_axis(lwpolylines):
    """
    Mirrors all points around the x-axis and makes all y-values positive.
    """
    if not isinstance(lwpolylines, list) or not all(isinstance(polyline, dict) and "points" in polyline for polyline in lwpolylines):
        raise ValueError("Input must be a list of dictionaries with a 'points' key.")

    if not lwpolylines:
        return lwpolylines  # Return empty list if input is empty

    # Find the maximum y value across all points
    max_y = float('-inf')
    for polyline in lwpolylines:
        for point in polyline["points"]:
            max_y = max(max_y, point[1])

    print(f"Maximum y value: {max_y}")

    # Mirror points across x-axis and make y positive
    for polyline in lwpolylines:
        updated_points = [
            (point[0], abs(point[1] - max_y)) for point in polyline["points"]
        ]
        polyline["points"] = updated_points

    return lwpolylines


# Function to convert numeric column index to Excel-style column label
def column_index_to_label(index):
    """
    Converts a numeric column index (0-based) to an Excel-style column label.
    For example, 0 -> 'A', 25 -> 'Z', 26 -> 'AA'.
    """
    label = ""
    while index >= 0:
        label = chr(index % 26 + ord('A')) + label
        index = index // 26 - 1
    return label

# Function to find the grid cell for a given center and orientation
def find_grid_cell(cellcenter, grid_origin, cell_width, cell_height, grid_rows, grid_columns, orientation):
    """
    Finds the grid cell where the given cellcenter falls, and also returns the adjacent cell
    based on the orientation of the rectangle (horizontal or vertical).
    """
    x_center, y_center = cellcenter
    x_origin, y_origin = grid_origin

    # Determine column and row indices based on the cellcenter
    col_idx = int((x_center - x_origin) // cell_width)
    row_idx = int((y_center - y_origin) // cell_height)

    # Check if the indices are within the grid bounds
    if 0 <= col_idx < grid_columns and 0 <= row_idx < grid_rows:
        # Get the column label using the column index
        column_label = column_index_to_label(col_idx)
        # Excel-style row is 1-based (e.g., 1 for the first row, 2 for the second row)
        row_label = row_idx + 1  # Row starts at 1, not 0

        # Current cell is the one where the cellcenter falls
        current_cell = f"{column_label}{row_label}"
        
        # Based on the orientation, find the adjacent cell
        if orientation == "Horizontal":
            # For horizontal rectangles, the adjacent cell is directly below
            if row_idx + 1 < grid_rows:
                adjacent_column_label = column_index_to_label(col_idx + 1)
                adjacent_cell = f"{adjacent_column_label}{row_label}"
            else:
                adjacent_cell = None  # No cell below if it's at the edge of the grid
        elif orientation == "Vertical":
            # For vertical rectangles, the adjacent cell is directly to the right
            if col_idx + 1 < grid_columns:
                adjacent_cell = f"{column_label}{row_label + 1}"
            else:
                adjacent_cell = None  # No cell to the right if it's at the edge of the grid
        
        return current_cell, adjacent_cell
    else:
        return None, None  # Cellcenter is out of grid bounds



# Process the polyline data and write to the Excel file
def process_polylines_to_excel(lwpolylines, grid_origin, cell_width, cell_height, grid_rows, grid_columns):

    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Grid Cells"

    # Set a uniform column width and row height
    uniform_size = 15  # Adjust as needed
    for col in range(1, 2000):  # Assuming up to column AX (adjust range if needed)
        ws.column_dimensions[get_column_letter(col)].width = 3
    for row in range(1, 2000):  # Assuming up to row 100 (adjust range if needed)
        ws.row_dimensions[row].height = 14.5

    # Define styles
    light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for idx, polyline in enumerate(lwpolylines, start=1):
        properties = check_rectangle_properties(polyline)
        
        if properties:
            orientation, center, avg_height, avg_width, cellcenter = properties
            grid_cell = find_grid_cell(cellcenter, grid_origin, cell_width, cell_height, grid_rows, grid_columns, orientation)
            
            if grid_cell:
                cell1, cell2 = grid_cell  # Unpack the two adjacent cell references
                
                # Merge the two cells
                ws.merge_cells(f"{cell1}:{cell2}")
                merged_cell = ws[cell1]
                
                # Apply formatting
                merged_cell.fill = light_blue_fill
                merged_cell.alignment = Alignment(horizontal="center", vertical="center")
                #merged_cell.value = f"{1}"  # Label merged cell
                
                # Apply borders around the entire merged range
                start_col = "".join(filter(str.isalpha, cell1))  # Extract column letters
                start_row = int("".join(filter(str.isdigit, cell1)))  # Extract row number
                end_col = "".join(filter(str.isalpha, cell2))  # Extract column letters
                end_row = int("".join(filter(str.isdigit, cell2)))  # Extract row number

                # Convert column letters to numbers using column_index_from_string
                start_col_num = column_index_from_string(start_col)
                end_col_num = column_index_from_string(end_col)

                # Apply borders only to the outermost cells in the merged range
                for row in range(start_row, end_row + 1):
                    for col in range(start_col_num, end_col_num + 1):
                        cell = ws.cell(row=row, column=col)
                        cell.border = thin_border  # Apply border to every cell in the merged range

            else:
                print(f"LWPOLYLINE {idx}: Cellcenter {cellcenter} is out of grid bounds.")
        else:
            print(f"LWPOLYLINE {idx}: Not a rectangle")

    # Save to an Excel file
    wb.save("grid_cells_output.xlsx")
    print("Excel file saved as 'grid_cells_output.xlsx'")



def master_function(file_path, grid_origin, cell_width, cell_height, grid_rows, grid_columns):
    
    lwpolylines = read_lwpolylines_from_dxf(file_path)
    
    # Step 2: Filter for valid rectangles
    valid_rectangles = [polyline for polyline in lwpolylines if validate_rectangle(polyline)]
    
    # Step 3: Check rectangle properties (orientation, center, etc.)
    for idx, polyline in enumerate(valid_rectangles, start=1):
        orientation, center, avg_height, avg_width, cellcenter = check_rectangle_properties(polyline)
        print(f"LWPOLYLINE {idx}:")
        print(f"  Orientation: {orientation}")
        print(f"  Center: {center}")
        print(f"  Average Height: {avg_height}")
        print(f"  Average Width: {avg_width}")

        # Find grid cell for the center and adjacent cell
        current_cell, adjacent_cell = find_grid_cell(center, grid_origin, cell_width, cell_height, grid_rows, grid_columns, orientation)
        print(f"  Current Grid Cell: {current_cell}")
        if adjacent_cell:
            print(f"  Adjacent Grid Cell: {adjacent_cell}")
        else:
            print(f"  No adjacent cell (edge of grid)")

    # Step 4: Move all polylines to the origin
    moved_polylines = move_polylines_to_origin(lwpolylines)

    # Step 5: Mirror points across the x-axis
    flipped_polylines = mirror_points_across_x_axis(moved_polylines)

    # Step 6: Print moved and mirrored polylines
    for idx, polyline in enumerate(flipped_polylines, start=1):
        print(f"LWPOLYLINE {idx} after moving to origin and mirroring:")
        print(f"  Points: {polyline['points']}")

    # Step 7: Process polylines and write to Excel
    process_polylines_to_excel(flipped_polylines, grid_origin, cell_width, cell_height, grid_rows, grid_columns)

    # Return a success message
    return {"message": "File processed and Excel file saved as 'grid_cells_output.xlsx'"}
